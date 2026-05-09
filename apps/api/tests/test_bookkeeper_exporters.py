"""Round-trip tests for the AAHA bookkeeper export adapters.

For each of the 6 supported formats we run the same fixture (a week of
30 line items across all AAHA top-level groups, including a refund
row, a flagged-for-review row, and rows for two practice locations)
through the adapter and assert:

  - the bytes are produced (smoke test)
  - the file is parseable by the canonical importer / spec for that
    format (schema validity)
  - format-specific invariants hold (Sage CSV journal-entry balance,
    QBO date format, IIF header sequence, Xero column stars, etc.)

These are pure-function tests — no DB, no API, no MCP. The adapters
take dataclasses in and return bytes out, which is exactly the round-
trip surface the format spec rejects on.
"""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import date

import pytest

from app.services.bookkeeper_exporters import (
    AAHATaxonomy,
    AAHATaxonomyLeaf,
    CategorizedLineItem,
    SUPPORTED_FORMATS,
    TenantExportMetadata,
    get_adapter,
)
from app.services.bookkeeper_exporters.csv import CSV_COLUMNS
from app.services.bookkeeper_exporters.quickbooks_iif import (
    SPL_HEADER,
    TRNS_HEADER,
)
from app.services.bookkeeper_exporters.quickbooks_qbo import QBO_COLUMNS
from app.services.bookkeeper_exporters.sage_intacct_csv import SAGE_COLUMNS
from app.services.bookkeeper_exporters.xero_csv import XERO_COLUMNS


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def taxonomy() -> AAHATaxonomy:
    """Minimal AAHA taxonomy slice — every category referenced by the
    fixture line items appears here. Loading the full YAML is exercised
    separately in test_bookkeeper_export_service."""
    return AAHATaxonomy(
        leaves=(
            AAHATaxonomyLeaf(
                name="Drugs and medical supplies",
                gl_code="5210",
                top_level="COST_OF_SERVICES",
                description="Medical drugs, vaccines, surgical disposables",
            ),
            AAHATaxonomyLeaf(
                name="Laboratory - outside",
                gl_code="5310",
                top_level="COST_OF_SERVICES",
            ),
            AAHATaxonomyLeaf(
                name="Salaries - DVM",
                gl_code="6010",
                top_level="PERSONNEL",
            ),
            AAHATaxonomyLeaf(
                name="Rent",
                gl_code="7010",
                top_level="FACILITY",
            ),
            AAHATaxonomyLeaf(
                name="Utilities - electric",
                gl_code="7110",
                top_level="FACILITY",
            ),
            AAHATaxonomyLeaf(
                name="Marketing - digital",
                gl_code="8010",
                top_level="MARKETING",
            ),
            AAHATaxonomyLeaf(
                name="Office supplies",
                gl_code="8210",
                top_level="ADMINISTRATIVE",
            ),
            AAHATaxonomyLeaf(
                name="Professional services - DVM exam",
                gl_code="4100",
                top_level="REVENUE",
            ),
        )
    )


@pytest.fixture
def fixture_items() -> list[CategorizedLineItem]:
    """30 line items spanning a week, multiple AAHA categories and 2 locations."""
    items: list[CategorizedLineItem] = []
    base = date(2026, 5, 4)  # a Monday
    # 25 routine expense rows across categories
    rows = [
        ("Patterson Veterinary",   "Drugs and medical supplies",   "5210",   142.55, "Northgate"),
        ("Patterson Veterinary",   "Drugs and medical supplies",   "5210",   88.20,  "Northgate"),
        ("IDEXX Reference Labs",   "Laboratory - outside",         "5310",   312.10, "Northgate"),
        ("IDEXX Reference Labs",   "Laboratory - outside",         "5310",   274.40, "Southside"),
        ("MWI Animal Health",      "Drugs and medical supplies",   "5210",   523.99, "Northgate"),
        ("Henry Schein",           "Drugs and medical supplies",   "5210",   206.45, "Southside"),
        ("Pacific Power",          "Utilities - electric",         "7110",   384.12, "Northgate"),
        ("Pacific Power",          "Utilities - electric",         "7110",   205.66, "Southside"),
        ("Northgate Realty",       "Rent",                         "7010",   4200.00, "Northgate"),
        ("Southside Holdings",     "Rent",                         "7010",   3100.00, "Southside"),
        ("Google Ads",             "Marketing - digital",          "8010",   150.00, "Northgate"),
        ("Meta Ads",               "Marketing - digital",          "8010",   220.00, "Northgate"),
        ("Office Depot",           "Office supplies",              "8210",   62.10,  "Northgate"),
        ("Costco Business",        "Office supplies",              "8210",   118.42, "Southside"),
        ("Patterson Veterinary",   "Drugs and medical supplies",   "5210",   65.00,  "Southside"),
        ("Antech Diagnostics",     "Laboratory - outside",         "5310",   189.30, "Northgate"),
        ("Antech Diagnostics",     "Laboratory - outside",         "5310",   154.20, "Southside"),
        ("Henry Schein",           "Drugs and medical supplies",   "5210",   77.55,  "Northgate"),
        ("MWI Animal Health",      "Drugs and medical supplies",   "5210",   412.30, "Southside"),
        ("Office Depot",           "Office supplies",              "8210",   23.99,  "Northgate"),
        ("Comcast Business",       "Utilities - electric",         "7110",   189.99, "Northgate"),
        ("Google Ads",             "Marketing - digital",          "8010",   85.00,  "Southside"),
        ("Patterson Veterinary",   "Drugs and medical supplies",   "5210",   315.00, "Northgate"),
        ("IDEXX Reference Labs",   "Laboratory - outside",         "5310",   240.50, "Northgate"),
        ("MWI Animal Health",      "Drugs and medical supplies",   "5210",   85.10,  "Southside"),
    ]
    for i, (vendor, cat, gl, amt, loc) in enumerate(rows):
        items.append(
            CategorizedLineItem(
                txn_date=date(2026, 5, 4 + (i % 7)),
                vendor=vendor,
                amount=amt,
                aaha_category=cat,
                gl_code=gl,
                location=loc,
                confidence=0.95,
                flagged_for_review=False,
                source_email_id=f"gmail-msg-{i:04d}",
                memo=f"Auto-categorized line {i + 1}",
                reference=f"INV-{2000 + i}",
            )
        )

    # 1 refund row (negative amount)
    items.append(
        CategorizedLineItem(
            txn_date=date(2026, 5, 8),
            vendor="Patterson Veterinary",
            amount=-42.10,
            aaha_category="Drugs and medical supplies",
            gl_code="5210",
            location="Northgate",
            confidence=0.92,
            flagged_for_review=False,
            source_email_id="gmail-msg-refund-1",
            memo="Credit for damaged shipment",
            reference="CRN-9001",
        )
    )

    # 2 flagged-for-review rows (low confidence, new vendor)
    items.append(
        CategorizedLineItem(
            txn_date=date(2026, 5, 6),
            vendor="Random LLC",
            amount=99.00,
            aaha_category="Office supplies",
            gl_code="8210",
            location="Northgate",
            confidence=0.42,
            flagged_for_review=True,
            source_email_id="gmail-msg-flag-1",
            memo="New vendor — confidence below floor",
        )
    )
    items.append(
        CategorizedLineItem(
            txn_date=date(2026, 5, 7),
            vendor="Mystery Vendor Inc",
            amount=440.00,
            aaha_category="Marketing - digital",
            gl_code="8010",
            location="Southside",
            confidence=0.51,
            flagged_for_review=True,
            source_email_id="gmail-msg-flag-2",
            memo="No prior history",
        )
    )

    # 1 row with a single-quote vendor name (CSV escaping torture test)
    items.append(
        CategorizedLineItem(
            txn_date=date(2026, 5, 5),
            vendor='Joe\'s Pet Supply, "Premier" branch',
            amount=51.20,
            aaha_category="Drugs and medical supplies",
            gl_code="5210",
            location="Northgate",
            confidence=0.88,
            flagged_for_review=False,
            source_email_id="gmail-msg-quote-1",
            memo='Memo with, comma and "quotes"',
        )
    )

    # 1 row with a tab in the memo (IIF torture test — IIF rejects raw tabs)
    items.append(
        CategorizedLineItem(
            txn_date=date(2026, 5, 8),
            vendor="Acme Veterinary",
            amount=33.33,
            aaha_category="Office supplies",
            gl_code="8210",
            location="Northgate",
            confidence=0.96,
            flagged_for_review=False,
            source_email_id="gmail-msg-tab-1",
            memo="Memo with\ttab character",
        )
    )

    assert len(items) == 30, "fixture must have 30 line items per the plan"
    return items


@pytest.fixture
def tenant_meta() -> TenantExportMetadata:
    return TenantExportMetadata(
        tenant_id="7f632730-1a38-41f1-9f99-508d696dbcf1",
        practice_name="The Animal Doctor SOC",
        period_start=date(2026, 5, 4),
        period_end=date(2026, 5, 10),
        locations=("Northgate", "Southside"),
        base_currency="USD",
    )


# ---------------------------------------------------------------------------
# Cross-format smoke tests
# ---------------------------------------------------------------------------


def test_registry_lists_all_six_formats():
    assert set(SUPPORTED_FORMATS) == {
        "xlsx", "csv", "quickbooks_iif", "quickbooks_qbo",
        "xero_csv", "sage_intacct_csv",
    }


@pytest.mark.parametrize("fmt", SUPPORTED_FORMATS)
def test_each_adapter_produces_nonempty_output(fmt, fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter(fmt)
    result = adapter(fixture_items, taxonomy, tenant_meta)
    assert isinstance(result.content, bytes)
    assert len(result.content) > 0
    assert result.filename
    assert result.mime_type


@pytest.mark.parametrize("fmt", SUPPORTED_FORMATS)
def test_each_adapter_is_deterministic(fmt, fixture_items, taxonomy, tenant_meta):
    """Same input → same bytes. Round-trip golden-master safety net."""
    adapter = get_adapter(fmt)
    a = adapter(fixture_items, taxonomy, tenant_meta)
    b = adapter(fixture_items, taxonomy, tenant_meta)
    if fmt == "xlsx":
        # XLSX zip metadata embeds creation timestamps, so we can't
        # byte-compare directly. Instead verify the sheet names match.
        with zipfile.ZipFile(io.BytesIO(a.content)) as za, \
             zipfile.ZipFile(io.BytesIO(b.content)) as zb:
            assert sorted(za.namelist()) == sorted(zb.namelist())
    else:
        assert a.content == b.content


# ---------------------------------------------------------------------------
# Format-specific schema validation
# ---------------------------------------------------------------------------


def test_csv_schema_validates(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("csv")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    text_io = io.StringIO(result.content.decode("utf-8"))
    reader = csv.reader(text_io)
    header = next(reader)
    assert tuple(header) == CSV_COLUMNS
    rows = list(reader)
    assert len(rows) == len(fixture_items)
    # Every row has the right column count (CSV importer rejection vector)
    for row in rows:
        assert len(row) == len(CSV_COLUMNS)
    assert result.mime_type == "text/csv"
    assert result.filename.endswith(".csv")


def test_quickbooks_iif_header_block_first(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("quickbooks_iif")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    text = result.content.decode("utf-8")
    # IIF Desktop importer requires the !-prefixed headers BEFORE any
    # data rows; getting this wrong rejects the entire file silently.
    lines = text.splitlines()
    assert lines[0].startswith("!TRNS\t")
    assert lines[1].startswith("!SPL\t")
    assert lines[2] == "!ENDTRNS"
    assert tuple(lines[0].split("\t")) == TRNS_HEADER
    assert tuple(lines[1].split("\t")) == SPL_HEADER

    # CRLF line endings — Desktop import rejects LF-only files
    assert "\r\n" in text

    # No raw tabs inside data fields — would split a row mid-cell
    for line in lines[3:]:
        # Field count check: every TRNS / SPL row matches the header column count
        if line.startswith("TRNS\t"):
            assert len(line.split("\t")) == len(TRNS_HEADER)
        if line.startswith("SPL\t"):
            assert len(line.split("\t")) == len(SPL_HEADER)

    # Each line item produces one TRNS + one SPL + one ENDTRNS
    trns_count = sum(1 for line in lines if line.startswith("TRNS\t"))
    spl_count = sum(1 for line in lines if line.startswith("SPL\t"))
    endtrns_count = sum(1 for line in lines if line == "ENDTRNS")
    assert trns_count == len(fixture_items)
    assert spl_count == len(fixture_items)
    assert endtrns_count == len(fixture_items)


def test_quickbooks_qbo_columns_and_date_format(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("quickbooks_qbo")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.reader(io.StringIO(result.content.decode("utf-8")))
    header = next(reader)
    assert tuple(header) == QBO_COLUMNS
    rows = list(reader)
    assert len(rows) == len(fixture_items)
    # QBO US-locale date format: MM/DD/YYYY
    for row in rows:
        assert len(row[0]) == 10
        assert row[0][2] == "/" and row[0][5] == "/"


def test_xero_csv_required_columns_have_stars(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("xero_csv")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.reader(io.StringIO(result.content.decode("utf-8")))
    header = next(reader)
    # Xero matches column names literally — getting the leading `*`
    # wrong silently mis-maps the import.
    assert header[0] == "*Date"
    assert header[1] == "*Amount"
    assert tuple(header) == XERO_COLUMNS

    rows = list(reader)
    assert len(rows) == len(fixture_items)

    # Sign convention flip — internal positive (expense) → Xero negative
    # (money out of the bank). Pick the first non-refund row.
    first_nonrefund = next(
        (it for it in fixture_items if it.amount > 0), fixture_items[0]
    )
    matching = [r for r in rows if r[2] == first_nonrefund.vendor]
    assert matching
    # Amount in Xero CSV is the bank-side amount = -our amount
    xero_amt = float(matching[0][1])
    assert xero_amt < 0


def test_sage_intacct_csv_journal_entries_balance(
    fixture_items, taxonomy, tenant_meta
):
    """Sage rejects journal entries where Debit total != Credit total
    per Batch Title. The adapter MUST emit two rows per line item that
    balance the journal entry. We verify per-batch DR/CR balance."""
    adapter = get_adapter("sage_intacct_csv")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.reader(io.StringIO(result.content.decode("utf-8")))
    header = next(reader)
    assert tuple(header) == SAGE_COLUMNS
    rows = list(reader)
    # 2 rows per line item — DR leg + CR leg
    assert len(rows) == 2 * len(fixture_items)

    # Group by Batch Title, verify per-batch DR sum == CR sum
    from collections import defaultdict
    batches: dict[str, dict[str, float]] = defaultdict(lambda: {"dr": 0.0, "cr": 0.0})
    for row in rows:
        batch_title = row[0]
        debit = float(row[5]) if row[5] else 0.0
        credit = float(row[6]) if row[6] else 0.0
        batches[batch_title]["dr"] += debit
        batches[batch_title]["cr"] += credit

    for batch_title, totals in batches.items():
        assert abs(totals["dr"] - totals["cr"]) < 0.01, (
            f"Sage batch {batch_title} unbalanced: "
            f"DR={totals['dr']} CR={totals['cr']}"
        )


def test_xlsx_has_expected_sheets(fixture_items, taxonomy, tenant_meta):
    """Smoke-test the XLSX bytes: open as zip (XLSX is a zipped XML
    bundle), assert the sheet structure matches the plan."""
    adapter = get_adapter("xlsx")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    assert result.mime_type.endswith(
        "spreadsheetml.sheet"
    )
    # XLSX = zip; opening as a zip is the canonical "did we produce a
    # valid file?" parse check, since the XLSX spec requires the zip be
    # readable and contain at least `xl/workbook.xml`.
    with zipfile.ZipFile(io.BytesIO(result.content)) as zf:
        names = zf.namelist()
        assert "xl/workbook.xml" in names

    # And via openpyxl we can verify the sheet titles
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(result.content), read_only=True)
    sheet_titles = wb.sheetnames
    # Required sheets per the plan
    assert "Cover" in sheet_titles
    assert "AAHA Categorized" in sheet_titles
    assert "Flagged for Review" in sheet_titles
    assert "Vendor Summary" in sheet_titles
    assert "AAHA Roll-up" in sheet_titles
    # Per-location tabs (this fixture has 2 locations)
    loc_sheets = [s for s in sheet_titles if s.startswith("Loc — ")]
    assert len(loc_sheets) == 2


# ---------------------------------------------------------------------------
# Data-integrity assertions (end-to-end "the numbers are right" checks)
# ---------------------------------------------------------------------------


def test_csv_amount_sum_matches_input(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("csv")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.DictReader(io.StringIO(result.content.decode("utf-8")))
    csv_total = sum(float(r["amount"]) for r in reader)
    expected = sum(it.amount for it in fixture_items)
    assert abs(csv_total - expected) < 0.01


def test_qbo_amount_sum_matches_input(fixture_items, taxonomy, tenant_meta):
    adapter = get_adapter("quickbooks_qbo")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.DictReader(io.StringIO(result.content.decode("utf-8")))
    qbo_total = sum(float(r["Amount"]) for r in reader)
    expected = sum(it.amount for it in fixture_items)
    assert abs(qbo_total - expected) < 0.01


def test_xero_amount_sum_is_inverted_input_sum(fixture_items, taxonomy, tenant_meta):
    """Xero positive = bank-deposit, our positive = expense → totals invert."""
    adapter = get_adapter("xero_csv")
    result = adapter(fixture_items, taxonomy, tenant_meta)
    reader = csv.DictReader(io.StringIO(result.content.decode("utf-8")))
    xero_total = sum(float(r["*Amount"]) for r in reader)
    expected = sum(it.amount for it in fixture_items)
    assert abs(xero_total + expected) < 0.01
