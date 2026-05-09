"""XLSX adapter — the AAHA-categorized weekly delivery sheet.

Sheet structure (per the 2026-05-09 plan):
  - "AAHA Categorized"      — every line item, all columns
  - one tab per location    — same shape filtered to that location
  - "Flagged for Review"    — rows that fell below the AAHA leaf's
                              confidence floor or are new-vendor first
                              sightings; the practice owner / Taylor
                              triages these before sign-off
  - "Vendor Summary"        — vendor → AAHA category → total amount
                              (quick visual catch-all for "is the
                              categorizer doing the right thing")
  - "AAHA Roll-up"          — top-level → leaf → amount; the page the
                              CPA actually scans

This adapter is built directly against `openpyxl` (already a dependency
via `apps/api/app/api/v1/reports.py`). Output is deterministic for a
given input — sheet order and row order are stable so round-trip
golden-master tests stay green.
"""

from __future__ import annotations

import io
from collections import defaultdict
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.bookkeeper_exporters.types import (
    AAHATaxonomy,
    CategorizedLineItem,
    ExportResult,
    TenantExportMetadata,
)


# ---------------------------------------------------------------------------
# Styling — kept in sync with `apps/api/app/api/v1/reports.py` so the
# Bookkeeper export feels visually identical to the existing operations
# report. CPAs glance at headers; consistent styling keeps muscle memory.
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
FLAG_FILL = PatternFill(start_color="FFEBCC", end_color="FFEBCC", fill_type="solid")

FMT_CURRENCY = '$#,##0.00'
FMT_PERCENT = '0.0%'

LINE_ITEM_HEADERS = (
    "Date",
    "Vendor",
    "Amount",
    "GL Code",
    "AAHA Category",
    "Top Level",
    "Location",
    "Confidence",
    "Flagged",
    "Source",
    "Memo",
    "Reference",
)


def _write_header_row(ws, row: int, values: Iterable[str]) -> int:
    for idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    return row + 1


def _write_line_item_row(
    ws,
    row: int,
    item: CategorizedLineItem,
    taxonomy: AAHATaxonomy,
) -> int:
    leaf = taxonomy.by_name(item.aaha_category)
    top_level = leaf.top_level if leaf else ""
    cells = [
        item.txn_date.isoformat(),
        item.vendor,
        round(item.amount, 2),
        item.gl_code,
        item.aaha_category,
        top_level,
        item.location,
        round(item.confidence, 4),
        "Yes" if item.flagged_for_review else "",
        item.source_email_id,
        item.memo,
        item.reference,
    ]
    for idx, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=idx, value=val)
        c.font = DATA_FONT
        if idx == 3:
            c.number_format = FMT_CURRENCY
            c.alignment = Alignment(horizontal="right")
        if idx == 8:
            c.number_format = FMT_PERCENT
        if item.flagged_for_review:
            c.fill = FLAG_FILL
    return row + 1


def _autosize(ws) -> None:
    """Cheap visual tidy — set sane column widths instead of measuring text."""
    widths = [12, 30, 14, 10, 28, 16, 18, 12, 9, 24, 32, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _build_line_items_sheet(
    wb: Workbook,
    title: str,
    items: list[CategorizedLineItem],
    taxonomy: AAHATaxonomy,
) -> None:
    ws = wb.create_sheet(title=title)
    row = _write_header_row(ws, 1, LINE_ITEM_HEADERS)
    for item in items:
        row = _write_line_item_row(ws, row, item, taxonomy)
    _autosize(ws)


def _build_vendor_summary(
    wb: Workbook,
    items: list[CategorizedLineItem],
) -> None:
    ws = wb.create_sheet(title="Vendor Summary")
    # vendor -> aaha_category -> total
    summary: dict[tuple[str, str], float] = defaultdict(float)
    for item in items:
        summary[(item.vendor, item.aaha_category)] += item.amount

    row = _write_header_row(ws, 1, ("Vendor", "AAHA Category", "Total Amount"))
    # Stable ordering: vendor name, then category name
    for (vendor, category), total in sorted(summary.items()):
        ws.cell(row=row, column=1, value=vendor).font = DATA_FONT
        ws.cell(row=row, column=2, value=category).font = DATA_FONT
        c = ws.cell(row=row, column=3, value=round(total, 2))
        c.number_format = FMT_CURRENCY
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16


def _build_aaha_rollup(
    wb: Workbook,
    items: list[CategorizedLineItem],
    taxonomy: AAHATaxonomy,
) -> None:
    ws = wb.create_sheet(title="AAHA Roll-up")
    # top_level -> leaf -> total
    rollup: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for item in items:
        leaf = taxonomy.by_name(item.aaha_category)
        top_level = leaf.top_level if leaf else "UNCLASSIFIED"
        rollup[top_level][item.aaha_category] += item.amount

    row = _write_header_row(ws, 1, ("Top Level", "AAHA Category", "GL Code", "Amount"))
    grand_total = 0.0
    for top_level in sorted(rollup):
        # top-level subheader row
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1, value=top_level).font = SUBHEADER_FONT
        row += 1
        subtotal = 0.0
        for cat in sorted(rollup[top_level]):
            leaf = taxonomy.by_name(cat)
            gl = leaf.gl_code if leaf else ""
            amt = round(rollup[top_level][cat], 2)
            ws.cell(row=row, column=1, value="").font = DATA_FONT
            ws.cell(row=row, column=2, value=cat).font = DATA_FONT
            ws.cell(row=row, column=3, value=gl).font = DATA_FONT
            c = ws.cell(row=row, column=4, value=amt)
            c.number_format = FMT_CURRENCY
            c.font = DATA_FONT
            c.alignment = Alignment(horizontal="right")
            subtotal += amt
            row += 1
        # subtotal
        ws.cell(row=row, column=2, value=f"  Subtotal {top_level}").font = SUBHEADER_FONT
        c = ws.cell(row=row, column=4, value=round(subtotal, 2))
        c.font = SUBHEADER_FONT
        c.number_format = FMT_CURRENCY
        c.alignment = Alignment(horizontal="right")
        row += 1
        grand_total += subtotal

    # Grand total
    ws.cell(row=row, column=2, value="GRAND TOTAL").font = TITLE_FONT
    c = ws.cell(row=row, column=4, value=round(grand_total, 2))
    c.font = TITLE_FONT
    c.number_format = FMT_CURRENCY
    c.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16


def _build_cover_sheet(
    wb: Workbook,
    tenant_metadata: TenantExportMetadata,
    items: list[CategorizedLineItem],
) -> None:
    ws = wb.active  # the workbook auto-creates one — repurpose it
    ws.title = "Cover"
    ws.cell(row=1, column=1, value=f"{tenant_metadata.practice_name} — Bookkeeper Export").font = TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Period: {tenant_metadata.period_start.isoformat()} — "
            f"{tenant_metadata.period_end.isoformat()}"
        ),
    ).font = DATA_FONT
    ws.cell(row=3, column=1, value="Format: AAHA-categorized (canonical)").font = DATA_FONT
    ws.cell(row=4, column=1, value=f"Currency: {tenant_metadata.base_currency}").font = DATA_FONT
    ws.cell(row=5, column=1, value=f"Total line items: {len(items)}").font = DATA_FONT
    flagged = sum(1 for i in items if i.flagged_for_review)
    ws.cell(row=6, column=1, value=f"Flagged for review: {flagged}").font = DATA_FONT
    ws.column_dimensions["A"].width = 60


def export(
    items: list[CategorizedLineItem],
    taxonomy: AAHATaxonomy,
    tenant_metadata: TenantExportMetadata,
) -> ExportResult:
    """Build the multi-tab AAHA workbook."""
    wb = Workbook()
    _build_cover_sheet(wb, tenant_metadata, items)

    _build_line_items_sheet(wb, "AAHA Categorized", list(items), taxonomy)

    # Per-location tabs (only when the tenant has >1 location). For
    # single-location tenants we skip — the AAHA Categorized tab already
    # covers everything.
    if len(tenant_metadata.locations) > 1:
        for loc in tenant_metadata.locations:
            loc_items = [i for i in items if i.location == loc]
            if not loc_items:
                continue
            # Excel sheet titles cap at 31 chars.
            title = f"Loc — {loc}"[:31]
            _build_line_items_sheet(wb, title, loc_items, taxonomy)

    flagged = [i for i in items if i.flagged_for_review]
    _build_line_items_sheet(wb, "Flagged for Review", flagged, taxonomy)

    _build_vendor_summary(wb, items)
    _build_aaha_rollup(wb, items, taxonomy)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    filename = (
        f"{tenant_metadata.safe_practice_slug()}_AAHA_"
        f"{tenant_metadata.period_slug()}.xlsx"
    )
    return ExportResult(
        content=content,
        filename=filename,
        mime_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        line_item_count=len(items),
        format="xlsx",
    )
